from openpyxl import load_workbook

# Load the workbook and select a worksheet
workbook = load_workbook(
    r"C:\Users\ShAmiR\OneDrive\Desktop\Patient.csv.xlsx")
sheet = workbook.active  # Or workbook['SheetName']

# Access cell values
i = 0
for row in sheet.iter_rows(values_only=True):
    if i != 0:
        print(str(row[2]).split(' ')[0],str(row[-2]).split(' ')[0])
        # ob = Patients()
        # ob.DOCTOR = Doctor.objects.order_by('?').first()
        # ob.ROOM = Room.objects.order_by('?').first()
        # ob.date = row[2]
        # ob.status = 'discharged'
        # ob.ddate = row[-2]
        # ob.name = row[1]
        # ob.number = row[5]
        # ob.gender = 'Male'
        # ob.dob = row[3]
        # ob.address = row[4]
        # ob.save()
    i = i + 1